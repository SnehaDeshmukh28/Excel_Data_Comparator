"""
File Comparator — Compare a new (primary) file against an old (secondary) file.
Supports: .xlsx, .xls, .csv

Logic:
  - "OK"            -> exact match in old file                        (no highlight)
  - "Updated Entry" -> key columns match, but other cells differ      (blue row + yellow changed cells)
  - "New Entry"     -> key columns not found in old file at all       (green row)
"""

import os
import sys
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import filedialog, font as tkfont

# ── Colours ───────────────────────────────────────────────────────────────────
LIGHT_BLUE_FILL   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
LIGHT_GREEN_FILL  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# ── UI Theme ──────────────────────────────────────────────────────────────────
BG        = "#0F1117"
CARD      = "#1A1D27"
BORDER    = "#2A2D3A"
ACCENT    = "#4F8EF7"
ACCENT2   = "#38E2A0"
DANGER    = "#F75F5F"
TEXT      = "#E8EAF0"
SUBTEXT   = "#7B7F96"
BTN_FG    = "#FFFFFF"
INPUT_BG  = "#252837"


# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def style_root(win, title="File Comparator", w=540, h=400):
    win.title(title)
    win.configure(bg=BG)
    win.resizable(False, False)
    cx = (win.winfo_screenwidth()  - w) // 2
    cy = (win.winfo_screenheight() - h) // 2
    win.geometry(f"{w}x{h}+{cx}+{cy}")


def label(parent, text, size=11, bold=False, color=TEXT, **kwargs):
    weight = "bold" if bold else "normal"
    return tk.Label(parent, text=text, bg=parent["bg"], fg=color,
                    font=("Segoe UI", size, weight), **kwargs)


def btn(parent, text, command, color=ACCENT, width=18, **kwargs):
    b = tk.Button(
        parent, text=text, command=command,
        bg=color, fg=BTN_FG, activebackground=color, activeforeground=BTN_FG,
        font=("Segoe UI", 10, "bold"), relief="flat", bd=0,
        width=width, cursor="hand2", padx=10, pady=8, **kwargs
    )
    b.bind("<Enter>", lambda e: b.config(bg=_lighten(color)))
    b.bind("<Leave>", lambda e: b.config(bg=color))
    return b


def _lighten(hex_color):
    """Return a slightly lighter version of a hex colour."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r, g, b = min(255, r + 30), min(255, g + 30), min(255, b + 30)
    return f"#{r:02x}{g:02x}{b:02x}"


def divider(parent):
    tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=20, pady=6)


def card_frame(parent, **kwargs):
    return tk.Frame(parent, bg=CARD, bd=0, highlightthickness=1,
                    highlightbackground=BORDER, **kwargs)


# ══════════════════════════════════════════════════════════════════════════════
# STEP WINDOWS
# ══════════════════════════════════════════════════════════════════════════════

class FileSelectorWindow(tk.Toplevel):
    """Step 1 & 2 — pick new + old files and output path."""

    def __init__(self, parent):
        super().__init__(parent)
        self.result      = None   # (new_file, old_file, output_file) or None
        self._new_path   = tk.StringVar()
        self._old_path   = tk.StringVar()
        self._out_path   = tk.StringVar()

        style_root(self, "File Comparator — Select Files", 560, 420)
        self.grab_set()
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=BG)
        hdr.pack(fill="x", padx=24, pady=(22, 4))
        label(hdr, "📂  File Comparator", size=16, bold=True, color=TEXT).pack(anchor="w")
        label(hdr, "Select the files to compare and where to save output.",
              size=9, color=SUBTEXT).pack(anchor="w", pady=(2, 0))

        divider(self)

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=24, pady=4)

        self._file_row(body, "New File  (primary)",   self._new_path, "#38E2A0")
        self._file_row(body, "Old File  (secondary)", self._old_path, "#4F8EF7")
        self._file_row(body, "Save Output As",        self._out_path, "#F7A84F", save=True)

        divider(self)

        foot = tk.Frame(self, bg=BG)
        foot.pack(fill="x", padx=24, pady=(0, 18))
        btn(foot, "Cancel", self._cancel, color="#3A3D50", width=10).pack(side="right", padx=(8, 0))
        btn(foot, "Next →", self._confirm, color=ACCENT, width=12).pack(side="right")

    def _file_row(self, parent, title, var, accent, save=False):
        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x", pady=6)
        label(row, title, size=9, bold=True, color=accent).pack(anchor="w")

        inp_row = tk.Frame(row, bg=BG)
        inp_row.pack(fill="x", pady=(3, 0))

        entry = tk.Entry(inp_row, textvariable=var, bg=INPUT_BG, fg=TEXT,
                         insertbackground=TEXT, relief="flat", font=("Segoe UI", 9),
                         bd=0, highlightthickness=1, highlightbackground=BORDER,
                         highlightcolor=accent)
        entry.pack(side="left", fill="x", expand=True, ipady=7, padx=(0, 8))

        cmd = (lambda v=var: self._pick_save(v)) if save else (lambda v=var: self._pick_open(v))
        tk.Button(inp_row, text="Browse", command=cmd,
                  bg=INPUT_BG, fg=accent, activebackground=BORDER, activeforeground=accent,
                  font=("Segoe UI", 9, "bold"), relief="flat", bd=0,
                  cursor="hand2", padx=12, pady=6).pack(side="left")

    def _pick_open(self, var):
        p = filedialog.askopenfilename(
            filetypes=[("Spreadsheet files", "*.xlsx *.xls *.csv"), ("All files", "*.*")])
        if p:
            var.set(p)

    def _pick_save(self, var):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="comparison_output.xlsx",
            filetypes=[("Excel file", "*.xlsx")])
        if p:
            var.set(p)

    def _confirm(self):
        new = self._new_path.get().strip()
        old = self._old_path.get().strip()
        out = self._out_path.get().strip()
        if not new or not old or not out:
            _toast(self, "Please fill in all three paths.")
            return
        for p in (new, old):
            if not os.path.exists(p):
                _toast(self, f"File not found:\n{p}")
                return
        self.result = (new, old, out)
        self.destroy()

    def _cancel(self):
        self.destroy()


class ColumnPickerWindow(tk.Toplevel):
    """Step 3 & 4 — pick key columns, optionally pick columns to skip."""

    def __init__(self, parent, columns: list):
        super().__init__(parent)
        self.columns     = columns
        self.key_cols    = []
        self.skip_cols   = []
        self._cancelled  = False

        style_root(self, "Configure Comparison", 560, 620)
        self.grab_set()
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=BG)
        hdr.pack(fill="x", padx=24, pady=(20, 4))
        label(hdr, "⚙️  Configure Comparison", size=15, bold=True).pack(anchor="w")
        label(hdr, "Choose key columns and optionally skip noisy columns.",
              size=9, color=SUBTEXT).pack(anchor="w", pady=(2, 0))

        divider(self)

        # Scrollable body
        canvas_frame = tk.Frame(self, bg=BG)
        canvas_frame.pack(fill="both", expand=True, padx=24, pady=4)

        canvas = tk.Canvas(canvas_frame, bg=BG, bd=0, highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        self._scroll_frame = tk.Frame(canvas, bg=BG)

        self._scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        def _on_mousewheel(e):
            try:
                canvas.yview_scroll(-1 * (e.delta // 120), "units")
            except Exception:
                pass
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        self.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Section A — Key columns
        self._key_vars  = {}
        self._skip_vars = {}
        self._build_section(
            "🔑  Key Columns  (required)",
            "These columns together uniquely identify each row.",
            ACCENT2, self._key_vars
        )

        tk.Frame(self._scroll_frame, bg=BG, height=10).pack()

        # Section B — Skip columns
        self._skip_label_frame = tk.Frame(self._scroll_frame, bg=BG)
        self._skip_label_frame.pack(fill="x", pady=(4, 0))
        label(self._skip_label_frame, "🚫  Skip Columns  (optional)",
              size=10, bold=True, color="#F7A84F").pack(anchor="w")
        label(self._skip_label_frame,
              "These columns will be ignored during comparison (e.g. last-updated timestamps).",
              size=9, color=SUBTEXT).pack(anchor="w", pady=(2, 6))

        self._skip_container = tk.Frame(self._scroll_frame, bg=BG)
        self._skip_container.pack(fill="x")

        self._skip_hidden = True
        self._build_skip_checkboxes()

        skip_btn_frame = tk.Frame(self._scroll_frame, bg=BG)
        skip_btn_frame.pack(fill="x", pady=(6, 0))
        self._toggle_skip_btn = tk.Button(
            skip_btn_frame,
            text="+ Show skip columns",
            command=self._toggle_skip,
            bg=BG, fg="#F7A84F", activebackground=BG, activeforeground="#F7A84F",
            font=("Segoe UI", 9, "bold"), relief="flat", bd=0, cursor="hand2"
        )
        self._toggle_skip_btn.pack(anchor="w")
        self._apply_skip_visibility()

        divider(self)

        foot = tk.Frame(self, bg=BG)
        foot.pack(fill="x", padx=24, pady=(0, 18))
        btn(foot, "Cancel", self._cancel, color="#3A3D50", width=10).pack(side="right", padx=(8, 0))
        btn(foot, "Run Comparison ▶", self._confirm, color=ACCENT2, width=18).pack(side="right")

    def _build_section(self, title, subtitle, accent, var_dict):
        sec = tk.Frame(self._scroll_frame, bg=BG)
        sec.pack(fill="x", pady=(4, 0))
        label(sec, title, size=10, bold=True, color=accent).pack(anchor="w")
        label(sec, subtitle, size=9, color=SUBTEXT).pack(anchor="w", pady=(2, 6))

        grid = tk.Frame(sec, bg=BG)
        grid.pack(fill="x")
        for i, col in enumerate(self.columns):
            var = tk.BooleanVar(value=False)
            var_dict[col] = var
            cb = tk.Checkbutton(
                grid, text=col, variable=var,
                bg=BG, fg=TEXT, selectcolor=INPUT_BG,
                activebackground=BG, activeforeground=accent,
                font=("Segoe UI", 9), anchor="w", cursor="hand2"
            )
            cb.grid(row=i // 2, column=i % 2, sticky="w", padx=(0, 20), pady=1)

    def _build_skip_checkboxes(self):
        for widget in self._skip_container.winfo_children():
            widget.destroy()
        grid = tk.Frame(self._skip_container, bg=BG)
        grid.pack(fill="x")
        for i, col in enumerate(self.columns):
            var = tk.BooleanVar(value=False)
            self._skip_vars[col] = var
            cb = tk.Checkbutton(
                grid, text=col, variable=var,
                bg=BG, fg=TEXT, selectcolor=INPUT_BG,
                activebackground=BG, activeforeground="#F7A84F",
                font=("Segoe UI", 9), anchor="w", cursor="hand2"
            )
            cb.grid(row=i // 2, column=i % 2, sticky="w", padx=(0, 20), pady=1)

    def _toggle_skip(self):
        self._skip_hidden = not self._skip_hidden
        self._apply_skip_visibility()

    def _apply_skip_visibility(self):
        if self._skip_hidden:
            self._skip_container.pack_forget()
            self._toggle_skip_btn.config(text="+ Show skip columns")
        else:
            self._skip_container.pack(fill="x")
            self._toggle_skip_btn.config(text="− Hide skip columns")

    def _confirm(self):
        key_cols  = [c for c, v in self._key_vars.items()  if v.get()]
        skip_cols = [c for c, v in self._skip_vars.items() if v.get()]

        if not key_cols:
            _toast(self, "Please select at least one key column.")
            return

        overlap = set(key_cols) & set(skip_cols)
        if overlap:
            _toast(self, f"Column(s) can't be both key and skip:\n{', '.join(overlap)}")
            return

        self.key_cols  = key_cols
        self.skip_cols = skip_cols
        self.destroy()

    def _cancel(self):
        self._cancelled = True
        self.destroy()


class ResultWindow(tk.Toplevel):
    """Final result popup."""

    def __init__(self, parent, ok, updated, new, output_file):
        super().__init__(parent)
        style_root(self, "Comparison Complete", 460, 320)
        self.grab_set()
        self._build(ok, updated, new, output_file)
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build(self, ok, updated, new, output_file):
        tk.Frame(self, bg=BG, height=12).pack()
        label(self, "✅  Comparison Complete", size=15, bold=True, color=ACCENT2).pack(pady=(8, 2))
        divider(self)

        stats = tk.Frame(self, bg=BG)
        stats.pack(pady=8)
        self._stat_row(stats, "✅  OK  (exact match)",    ok,      TEXT)
        self._stat_row(stats, "🔵  Updated Entry",        updated, "#4F8EF7")
        self._stat_row(stats, "🟢  New Entry",            new,     "#38E2A0")

        divider(self)

        out_frame = tk.Frame(self, bg=BG)
        out_frame.pack(fill="x", padx=24, pady=4)
        label(out_frame, "Output saved to:", size=9, color=SUBTEXT).pack(anchor="w")
        label(out_frame, output_file, size=8, color=TEXT, wraplength=400, justify="left").pack(anchor="w")

        tk.Frame(self, bg=BG, height=8).pack()
        btn(self, "Close", self.destroy, color=ACCENT, width=14).pack()
        tk.Frame(self, bg=BG, height=14).pack()

    def _stat_row(self, parent, label_text, count, color):
        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x", pady=3, padx=40)
        label(row, label_text, size=10, color=color).pack(side="left")
        label(row, str(count), size=10, bold=True, color=color).pack(side="right")


class ErrorWindow(tk.Toplevel):
    def __init__(self, parent, title, message):
        super().__init__(parent)
        style_root(self, title, 460, 220)
        self.grab_set()
        tk.Frame(self, bg=BG, height=16).pack()
        label(self, "⚠️  " + title, size=13, bold=True, color=DANGER).pack()
        divider(self)
        label(self, message, size=9, color=TEXT, wraplength=400, justify="left").pack(padx=24, pady=8)
        btn(self, "OK", self.destroy, color=DANGER, width=10).pack(pady=8)
        self.protocol("WM_DELETE_WINDOW", self.destroy)


def _toast(parent, msg):
    """Inline error toast — small popup near the parent window."""
    t = tk.Toplevel(parent)
    t.overrideredirect(True)
    t.configure(bg=DANGER)
    px, py = parent.winfo_rootx(), parent.winfo_rooty()
    pw, ph = parent.winfo_width(), parent.winfo_height()
    tw, th = 320, 52
    t.geometry(f"{tw}x{th}+{px + (pw - tw)//2}+{py + ph//2}")
    tk.Label(t, text=msg, bg=DANGER, fg="white",
             font=("Segoe UI", 9), wraplength=300, justify="center").pack(expand=True)
    t.after(2400, t.destroy)


# ══════════════════════════════════════════════════════════════════════════════
# FILE I/O
# ══════════════════════════════════════════════════════════════════════════════

def read_file(path: str) -> pd.DataFrame:
    print(f"[read_file] Reading: {path}")
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(path, dtype=str)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(path, dtype=str)
        else:
            raise ValueError(f"Unsupported format '{ext}'. Use .csv, .xlsx, or .xls.")
        df = df.fillna("")
        print(f"[read_file] Loaded {len(df)} rows, {len(df.columns)} cols.")
        return df
    except ValueError:
        raise
    except Exception as e:
        raise RuntimeError(f"Failed to read '{path}': {e}") from e


def validate_headers(new_df, old_df):
    print("[validate_headers] Comparing headers...")
    nh, oh = list(new_df.columns), list(old_df.columns)
    if nh == oh:
        print(f"[validate_headers] OK — {len(nh)} columns match.")
        return
    issues = []
    if len(nh) != len(oh):
        issues.append(f"  Column count: new={len(nh)}, old={len(oh)}")
    ns, os_ = set(nh), set(oh)
    missing = os_ - ns
    extra   = ns - os_
    if missing: issues.append(f"  In OLD not NEW: {sorted(missing)}")
    if extra:   issues.append(f"  In NEW not OLD: {sorted(extra)}")
    if not missing and not extra and nh != oh:
        issues.append(f"  Column order differs.\n  NEW:{nh}\n  OLD:{oh}")
    raise ValueError("Header validation failed:\n" + "\n".join(issues))


# ══════════════════════════════════════════════════════════════════════════════
# CORE COMPARISON
# ══════════════════════════════════════════════════════════════════════════════

def make_key(row, key_cols):
    return tuple(str(row[c]).strip() for c in key_cols)


def differing_col_indices(new_row, old_row, skip_indices: set) -> list:
    try:
        nv = [str(v).strip() for v in new_row]
        ov = [str(v).strip() for v in old_row]
        return [i for i, (n, o) in enumerate(zip(nv, ov))
                if n != o and i not in skip_indices]
    except Exception as e:
        print(f"[differing_col_indices] WARNING: {e}")
        return []


def compare(new_file, old_file, output_file, key_cols, skip_cols):
    print("\n" + "=" * 60)
    print(f"[compare] New    : {new_file}")
    print(f"[compare] Old    : {old_file}")
    print(f"[compare] Output : {output_file}")
    print(f"[compare] Keys   : {key_cols}")
    print(f"[compare] Skip   : {skip_cols}")
    print("=" * 60)

    new_df = read_file(new_file)
    old_df = read_file(old_file)
    validate_headers(new_df, old_df)

    if new_df.empty:
        print("[compare] WARNING: New file is empty.")
        new_df["compare_comments"] = []
        new_df.to_excel(output_file, index=False)
        return 0, 0, 0

    # Build skip index set (0-based col positions)
    all_cols    = list(new_df.columns)
    skip_indices = {all_cols.index(c) for c in skip_cols if c in all_cols}
    print(f"[compare] Skip column indices: {skip_indices}")

    # Build old-file key lookup: key -> row Series
    print("[compare] Building old-file lookup...")
    old_key_map = {}
    for i, (_, row) in enumerate(old_df.iterrows()):
        try:
            old_key_map[make_key(row, key_cols)] = row
        except Exception as e:
            print(f"[compare] WARNING: Could not index old row {i}: {e}")
    print(f"[compare] Lookup ready: {len(old_key_map)} keys.")

    # Build full exact-row set (excluding skip cols)
    old_full_set = set()
    for _, row in old_df.iterrows():
        try:
            vals = tuple(str(v).strip() for i, v in enumerate(row) if i not in skip_indices)
            old_full_set.add(vals)
        except Exception:
            pass

    comments, green_rows, blue_rows, yellow_cells = [], [], [], {}

    print(f"[compare] Processing {len(new_df)} rows...")
    for idx, row in new_df.iterrows():
        try:
            # Full match (ignoring skip cols)?
            full_key = tuple(str(v).strip() for i, v in enumerate(row) if i not in skip_indices)
            if full_key in old_full_set:
                comments.append("OK")
                continue

            k = make_key(row, key_cols)
            if k not in old_key_map:
                comments.append("New Entry")
                green_rows.append(idx)
                print(f"[compare] Row {idx:>4}: New Entry     [GREEN]")
            else:
                old_row   = old_key_map[k]
                diff_cols = differing_col_indices(row, old_row, skip_indices)
                changed   = [all_cols[c] for c in diff_cols if c < len(all_cols)]
                comments.append("Updated Entry")
                blue_rows.append(idx)
                yellow_cells[idx] = diff_cols
                print(f"[compare] Row {idx:>4}: Updated Entry [BLUE+YELLOW]  changed: {changed}")

        except Exception as e:
            print(f"[compare] ERROR on row {idx}: {e}\n{traceback.format_exc()}")
            comments.append("Error")

    while len(comments) < len(new_df):
        comments.append("Error")

    new_df["compare_comments"] = comments

    # Write Excel
    print(f"\n[compare] Writing: {output_file}")
    try:
        new_df.to_excel(output_file, index=False)
    except PermissionError:
        raise RuntimeError(f"'{output_file}' is open in Excel. Close it and retry.")
    except Exception as e:
        raise RuntimeError(f"Failed to write output: {e}") from e

    # Highlights
    print("[compare] Applying highlights...")
    try:
        wb, ws  = load_workbook(output_file), None
        wb_ws   = wb.active
        ws      = wb_ws
        ncols   = ws.max_column

        for df_idx in green_rows:
            er = df_idx + 2
            for c in range(1, ncols + 1):
                ws.cell(row=er, column=c).fill = LIGHT_GREEN_FILL
        print(f"[compare] Green  → {len(green_rows)} row(s).")

        for df_idx in blue_rows:
            er       = df_idx + 2
            diff_set = set(yellow_cells.get(df_idx, []))
            for c in range(1, ncols + 1):
                ws.cell(row=er, column=c).fill = (
                    LIGHT_YELLOW_FILL if (c - 1) in diff_set else LIGHT_BLUE_FILL
                )
        print(f"[compare] Blue+Yellow → {len(blue_rows)} row(s).")

        for col_cells in ws.columns:
            try:
                ml = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 4, 60)
            except Exception:
                pass

        wb.save(output_file)
        print(f"[compare] Saved: {output_file}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Failed to apply highlights: {e}") from e

    total   = len(new_df)
    ok      = total - len(green_rows) - len(blue_rows)
    updated = len(blue_rows)
    new     = len(green_rows)

    print("\n" + "=" * 60)
    print("SUMMARY")
    print(f"  Total          : {total}")
    print(f"  OK             : {ok}")
    print(f"  Updated Entry  : {updated}")
    print(f"  New Entry      : {new}")
    print(f"  Output         : {output_file}")
    print("=" * 60 + "\n")

    return ok, updated, new


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("[main] Starting File Comparator...")

    try:
        root = tk.Tk()
        root.withdraw()
    except Exception as e:
        print(f"[main] ERROR: tkinter init failed: {e}")
        sys.exit(1)

    # Step 1+2+3 — pick files
    file_win = FileSelectorWindow(root)
    root.wait_window(file_win)

    if not file_win.result:
        print("[main] File selection cancelled.")
        os._exit(0)

    new_file, old_file, output_file = file_win.result

    # Read columns for the picker
    try:
        sample = read_file(new_file)
        columns = list(sample.columns)
    except Exception as e:
        ErrorWindow(root, "Read Error", str(e))
        root.wait_window()
        os._exit(1)

    # Step 3+4 — pick key cols + skip cols
    col_win = ColumnPickerWindow(root, columns)
    root.wait_window(col_win)

    if col_win._cancelled:
        print("[main] Column selection cancelled.")
        os._exit(0)

    key_cols  = col_win.key_cols
    skip_cols = col_win.skip_cols

    # Run comparison
    try:
        ok, updated, new = compare(new_file, old_file, output_file, key_cols, skip_cols)
        res_win = ResultWindow(root, ok, updated, new, output_file)
        root.wait_window(res_win)

    except ValueError as e:
        print(f"[main] Validation error: {e}")
        ew = ErrorWindow(root, "Header Mismatch", str(e))
        root.wait_window(ew)

    except RuntimeError as e:
        print(f"[main] Runtime error: {e}")
        ew = ErrorWindow(root, "Error", str(e))
        root.wait_window(ew)

    except Exception as e:
        tb = traceback.format_exc()
        print(f"[main] Unexpected: {tb}")
        ew = ErrorWindow(root, "Unexpected Error", f"{e}\n\nSee console for details.")
        root.wait_window(ew)

    finally:
        print("[main] Shutting down.")
        try:
            root.destroy()
        except Exception:
            pass
        os._exit(0)


if __name__ == "__main__":
    main()